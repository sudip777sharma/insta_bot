from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from pynput.mouse import Button, Controller
import openpyxl
class InstaBot:
    def __init__(self, username, pw):
        self.username = username
        # self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(executable_path='C:\Python\Python37\Lib\site-packages\chromedriver.exe')
        self.driver.maximize_window()
        self.mouse = Controller()
        self.mouse.position = (1341, 91)
        self.mouse.click(Button.left, 1)
        self.driver.get("https://instagram.com")
        sleep(1)
        self.driver.find_element_by_xpath("//input[@name=\"username\"]") \
            .send_keys(username)
        self.driver.find_element_by_xpath("//input[@name=\"password\"]") \
            .send_keys(pw)
        self.driver.find_element_by_xpath('//button[@type="submit"]') \
            .click()
        sleep(8)
        self.driver.find_element_by_xpath("//button[contains(text(),'Not Now')]") \
            .click()

    def get_unfollowers(self):
        self.driver.find_element_by_xpath("//a[contains(@href,'/{}')]".format(self.username)) \
            .click()
        sleep(1)
        self.driver.find_element_by_xpath("//a[contains(@href,'/sudeep777sharma/following/')]") \
            .click()
        sleep(1)
        scroll_box = self.driver.find_element_by_xpath("/html/body/div[4]/div/div[2]")
        last_ht, ht = 0, 1
        while last_ht != ht:
            last_ht = ht
            sleep(1)
            ht = self.driver.execute_script("""
                arguments[0].scrollTo(0 , arguments[0].scrollHeight);
                return arguments[0].scrollHeight;
                """, scroll_box)

        links = scroll_box.find_elements_by_tag_name('a')
        names = [name.text for name in links if name.text != '']
        print(names)
        self.driver.find_element_by_xpath("/html/body/div[4]/div/div[1]/div/div[2]/button") \
            .click()

        self.driver.find_element_by_xpath("//a[contains(@href,'/sudeep777sharma/followers/')]") \
            .click()
        sleep(1)
        scroll_box = self.driver.find_element_by_xpath("/html/body/div[4]/div/div[2]")
        last_ht, ht = 0, 1
        while last_ht != ht:
            last_ht = ht
            sleep(1)
            ht = self.driver.execute_script("""
                arguments[0].scrollTo(0 , arguments[0].scrollHeight);
                return arguments[0].scrollHeight;
                """, scroll_box)

        links1 = scroll_box.find_elements_by_tag_name('a')
        names1 = [name.text for name in links1 if name.text != '']
        print(names1)
        self.driver.find_element_by_xpath("/html/body/div[4]/div/div[1]/div/div[2]/button").click()

    def surf(self):
        i = 1
        pix = 500
        while i < 1000000:
            i = i + 1
            if pix == i * 10:
                pix = pix + 500
                sleep(2)
            self.driver.execute_script("window.scrollBy(0,15)", "")

    def like_follower(self):
        sleep(3)
        self.driver.find_element_by_xpath("//input[@placeholder=\"Search\"]")\
            .send_keys("cutest_bad_boy7")
        sleep(3)
        self.driver.find_element_by_xpath("//*[@id='react-root']/section/nav/div[2]/div/div/div[2]/div[2]/div[2]/div/a[1]/div/div[1]").click()

path = "C:\jfile\insta_ids.xlsm"
wb = openpyxl.load_workbook(path)
sh = wb.active
rows = sh.max_row
column = sh.max_column
for r in range(1, rows+1):
    print(sh.cell(row=r, column=1).value, " , ", sh.cell(row=r,  column=2).value)
    my_bot = InstaBot(sh.cell(row=r, column=1).value, sh.cell(row=r,  column=2).value)
    #my_bot.surf()
    my_bot.like_follower()